#!/usr/bin/env python
#
# BitCurator
# 
# This code is distributed under the terms of the GNU General Public 
# License, Version 3. See the text file "COPYING" for further details 
# about the terms of this license.
#
# Generate report in Excel format (from xml input)
# 

import sys,os,shelve
import re,dfxml,fiwalk
from bc_utils import filename_from_path

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

from openpyxl.utils import get_column_letter

def build_local_wb(ws, fi, row_idx):
    ws.cell(row=row_idx, column=1,  value=fi.partition())
    ws.cell(row=row_idx, column=2,  value=fi.filename())
    ws.cell(row=row_idx, column=3,  value=fi.ext())
    ws.cell(row=row_idx, column=4,  value=str(fi.filesize()))
    ws.cell(row=row_idx, column=5,  value=str(fi.libmagic()))
    ws.cell(row=row_idx, column=6,  value=str(fi.ctime()))
    ws.cell(row=row_idx, column=7,  value=str(fi.atime()))
    ws.cell(row=row_idx, column=8,  value=str(fi.crtime()))
    ws.cell(row=row_idx, column=9,  value=str(fi.mtime()))
    ws.cell(row=row_idx, column=10, value=str(fi.md5()))
    ws.cell(row=row_idx, column=11, value=str(fi.sha1()))

def process_files(fn, ws):

    row_idx = [2]

    # Callback function to process the SAX stream
    def cb(fi):
        # Build a row for all regular files
        if fi.is_file():
            build_local_wb(ws, fi, row_idx[0])
            row_idx[0] += 1
        # Certain HFS volumes may have a "-" name_type. Check and continue:
        elif fi.name_type() == '-' and fi.meta_type == '1':
            build_local_wb(ws, fi, row_idx[0])
            row_idx[0] += 1

    fiwalk.fiwalk_using_sax(xmlfile=open(fn, 'rb'),callback=cb)

def bc_generate_xlsx(fn):

    wb = Workbook()
    #wb = Workbook(optimized_write = True)
    dest_filename = fn.outdir + "/" + filename_from_path(fn.fiwalk_xmlfile) + ".xlsx"
    print("Generating Excel report ", dest_filename)
    ws = wb.worksheets[0]
    ws.title = "File Object Information"

    ws['A1'] = "Partition"
    ws['B1'] = "Filename"
    ws['C1'] = "Extension"
    ws['D1'] = "Filesize"
    ws['E1'] = "File format"
    ws['F1'] = "Change time"
    ws['G1'] = "Access time"
    ws['H1'] = "Create time"
    ws['I1'] = "Modification time"
    ws['J1'] = "MD5 Hash"
    ws['K1'] = "SHA1 Hash"

    process_files(fn.fiwalk_xmlfile, ws)

    # Save the workbook to the open file
    wb.save(filename=dest_filename)


